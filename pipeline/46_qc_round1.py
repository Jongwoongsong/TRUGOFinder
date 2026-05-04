"""
46_qc_round1.py
Conservative QC pass on GOF HC dataset (2,480 rows) after targeted validation review.

Input:
  data/gof_human_curated_mechanism_review.xlsx  (Curated_All)
  data/gof_extraction_benchmark_with_targeted_external_tools.xlsx  (GROBID quality)

Output:
  data/gof_curated_after_manual_qc_round1.xlsx

QC rules applied (conservative — original rows never deleted, only labelled):
  A. EXCLUDE_NONHUMAN_EVIDENCE   — evidence text clearly from non-human organism
  B. EXCLUDE_GENE_EVIDENCE_MISMATCH — gene ≠ evidence gene (LOXL2/loricrin)
  C. REVIEW_POSSIBLE_LOF_HAPLOINSUFFICIENCY — OPA1 PMID 14961560 (variant table, no GOF cue)
  D. EXCLUDE_NO_GOF_EVIDENCE     — explicit "no gain-of-function demonstrated" in evidence
  E. EXCLUDE_NON_GOF_LOF (EMD)  — X-linked EMD, haploinsufficiency, no GOF override
  F. EXCLUDE_NON_GOF_LOF (other specific genes) — BSG, CFHR5, TBX6, CLCN2, LRRC8A
  G. EXCLUDE_NON_GOF_LOF (KCNJ1 specific PMIDs) — locked-closed / NOT_GOF
  H. REVIEW_MECHANISM             — conflicting GOF + no-GOF signals
  I. REVIEW_PMID_MISSING_POSSIBLE_GOF — KAT6A, ALAS2 without PMID
  J. ADD                          — new candidates from Candidate_Additions review
"""

import re
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment

HERE = Path(__file__).resolve().parent.parent
IN_XL   = HERE / "data/outputs/gof_human_curated_mechanism_review.xlsx"
TV_XL   = HERE / "data/outputs/gof_extraction_benchmark_with_targeted_external_tools.xlsx"
OUT_XL  = HERE / "data/outputs/gof_curated_after_manual_qc_round1.xlsx"

# ── 1. Load data ───────────────────────────────────────────────────────────────
df = pd.read_excel(IN_XL, sheet_name="Curated_All", dtype=str).fillna("")
print(f"Loaded {len(df)} rows from Curated_All")

df_tv = pd.read_excel(TV_XL, sheet_name="Targeted_Row_Validation", dtype=str).fillna("")
print(f"Loaded {len(df_tv)} rows from Targeted_Row_Validation")

# ── 2. Regex patterns ──────────────────────────────────────────────────────────
NONHUMAN_RE = re.compile(
    # Rice/plant — specific enough (GS3 = rice Grain Size 3; Oryza = rice genus)
    r"\bGS3\b|Grain Size 3|\bOryza\b|oryza sativa|"
    # Maize — haploid induction is a maize-specific phenotype; Zea mays is the species name
    r"haploid induction|\bZea mays\b|corn kernel|"
    # Daphnia — obligate parthenogenesis phrasing uniquely non-human
    r"\bDaphnia\b|obligately asexual|asexual clone|"
    # Drosophila-specific gene names (NOT just "Drosophila" which appears in human model papers)
    r"\bdCREB2\b|\bdCREB\b",
    # Deliberately excluded: \bDrosophila\b, \brice\b, \bNLD\b, melanogaster
    # — these words appear in human papers citing model organism validations
    re.I,
)

NEGATION_GOF_RE = re.compile(
    r"no gain.of.function effect demonstrated|"
    r"no GOF effect|"
    r"not gain.of.function|"
    r"did not demonstrate gain.of.function|"
    r"(?:lacks?|show\s+no).{0,25}gain.of.function|"
    r"no gain-of-function",
    re.I,
)

LOF_RE = re.compile(
    r"loss.of.function|loss-of-function|"
    r"\bhaploinsufficienc|\bhaploinsufficient\b|"
    r"\bnonfunctional\b|non.functional\b|"
    r"locked.closed conformation|locked-closed|"
    r"does not induce.{0,30}current|\bno current\b|"
    r"loss of MCT1|MCT1 loss|"
    r"drastically diminish.{0,20}VRAC|diminish.{0,20}VRAC activity|"
    r"emerin.{0,30}loss|loss.{0,30}emerin|"
    r"unable to encode.{0,30}full.length protein|"
    r"loss of functional.{0,20}channel|"
    r"reducing chloride efflux|"
    r"reduced serum.{0,30}level|"               # CFHR5: reduced serum CFHR5 levels
    r"decreased transcription activity|"          # TBX6: significantly decreased transcription activity
    r"significantly decreased.{0,30}activity",    # general decreased-activity LOF
    re.I,
)

GOF_OVERRIDE_RE = re.compile(
    r"constitutiv.{0,25}activ|constitutively active|"
    r"gain.of.function|gain of function|\bGOF\b|"
    r"dominant.negative|"
    r"neomorphic|"
    r"toxic gain|"
    r"hypermorph|"
    r"activating mutation",
    re.I,
)

LORICRIN_RE = re.compile(r"loricrin|\bLOR\b|mutant loricrin", re.I)

BAD_GROBID_SECTION_RE = re.compile(
    r"purchased from|taqman|pcr primer|biosystems|foster city|"
    r"fax:|tel\.|e-mail:|\w+@\w+\.(edu|com)|"
    r"department of.{0,40}medic|hospital,.{0,40}st\.|"
    r"harvard medical|institute for|university of",
    re.I,
)

# ── 3. GROBID quality lookup (gene, pmid, hgvs_coding) → quality string ────────
def _grobid_quality(row: pd.Series) -> str:
    if str(row.get("grobid_variant_found", "")).lower() != "true":
        return ""
    snippet = str(row.get("grobid_snippet", "") or row.get("grobid_variant_context", ""))
    if BAD_GROBID_SECTION_RE.search(snippet):
        return "BAD_SECTION"
    if str(row.get("grobid_gof_cue_found", "")).lower() == "true":
        return "GOOD"
    return "VARIANT_ONLY_NO_GOF"

gq_lookup: dict = {}
for _, r in df_tv.iterrows():
    key = (str(r.get("gene", "")), str(r.get("pmid", "")), str(r.get("hgvs_coding", "")))
    gq_lookup[key] = _grobid_quality(r)

# ── 4. QC classification function ─────────────────────────────────────────────
def classify_qc(row: pd.Series):
    """
    Returns (qc_round1_action, qc_round1_reason).
    Only makes deterministic, conservative calls — ambiguous rows → KEEP or REVIEW.
    """
    gene   = str(row.get("gene", "")).strip()
    pmid   = str(row.get("pmid", "")).strip()
    hgvs   = str(row.get("hgvs_coding", "")).strip()
    strict = str(row.get("strict_gof_status", "")).strip()
    gof_ef = str(row.get("gof_effect", ""))
    ev     = str(row.get("evidence_text", ""))
    title  = str(row.get("title", ""))
    combined = " ".join([gof_ef, ev, title])

    # ── A. Non-human organism evidence ──
    if NONHUMAN_RE.search(combined):
        return (
            "EXCLUDE_NONHUMAN_EVIDENCE",
            "Evidence/title contains non-human organism context (rice/maize/Drosophila/Daphnia/GS3/NLD/dCREB2)",
        )

    # ── B. Gene-evidence mismatch: LOXL2 rows whose evidence describes loricrin ──
    if gene == "LOXL2" and LORICRIN_RE.search(combined):
        return (
            "EXCLUDE_GENE_EVIDENCE_MISMATCH",
            "Gene=LOXL2 but evidence describes loricrin (LOR gene); likely extraction error",
        )

    # ── C. OPA1 PMID 14961560 — variant table, no GOF cue in fulltext ──
    if gene == "OPA1" and pmid == "14961560":
        return (
            "REVIEW_POSSIBLE_LOF_HAPLOINSUFFICIENCY",
            "OPA1 frameshift from PMID 14961560 — variant catalogue table; no GOF cue in fulltext (gof_cue=False)",
        )

    # ── D. Explicit negation of GOF in evidence/gof_effect ──
    if NEGATION_GOF_RE.search(combined):
        # Strip the negation phrase(s) from the text before checking for positive GOF override.
        # This prevents "no gain-of-function" from falsely triggering GOF_OVERRIDE_RE,
        # since "gain-of-function" is a substring of "no gain-of-function".
        combined_stripped = NEGATION_GOF_RE.sub(" ", combined)
        if GOF_OVERRIDE_RE.search(combined_stripped):
            return (
                "REVIEW_MECHANISM",
                "Conflicting: explicit no-GOF statement AND independent GOF keyword present — manual review needed",
            )
        return (
            "EXCLUDE_NO_GOF_EVIDENCE",
            "Explicit statement: 'no gain-of-function effect demonstrated' or equivalent",
        )

    # ── E. EMD: X-linked Emery-Dreifuss MD — haploinsufficiency/emerin loss ──
    if gene == "EMD":
        if GOF_OVERRIDE_RE.search(combined):
            return (
                "REVIEW_MECHANISM",
                "EMD: LOF disease gene (Emery-Dreifuss MD) but GOF override keyword present — review",
            )
        return (
            "EXCLUDE_NON_GOF_LOF",
            "EMD: X-linked Emery-Dreifuss MD — emerin loss/haploinsufficiency, no GOF evidence",
        )

    # ── F. Specific LOF-only genes: BSG, CFHR5, TBX6, CLCN2, LRRC8A ──
    LOF_SPECIFIC_GENES = {"BSG", "CFHR5", "TBX6", "CLCN2", "LRRC8A"}
    if gene in LOF_SPECIFIC_GENES:
        if LOF_RE.search(combined) and not GOF_OVERRIDE_RE.search(combined):
            return (
                "EXCLUDE_NON_GOF_LOF",
                f"{gene}: evidence describes LOF/functional reduction only; no GOF override present",
            )

    # ── G. KCNJ1 specific: locked-closed (10532965) or NOT_GOF (32590952) ──
    if gene == "KCNJ1":
        if pmid == "10532965" and not GOF_OVERRIDE_RE.search(combined):
            return (
                "EXCLUDE_NON_GOF_LOF",
                "KCNJ1 PMID 10532965: nonfunctional locked-closed channel conformation — mechanistic LOF",
            )
        if strict == "NOT_GOF":
            return (
                "EXCLUDE_NON_GOF_LOF",
                f"KCNJ1: strict_gof_status=NOT_GOF (loss-of-function frameshift); PMID {pmid}",
            )

    # ── H. PMID-missing special cases ──
    if not pmid:
        if gene == "KAT6A":
            return (
                "REVIEW_PMID_MISSING_POSSIBLE_GOF",
                "KAT6A: plausible GOF (KAT6A syndrome), but no PMID — cannot validate",
            )
        if gene == "ALAS2" and "1670_1671" in hgvs:
            return (
                "REVIEW_PMID_MISSING_POSSIBLE_GOF",
                "ALAS2 c.1670_1671TC>GA: 'increased specific activities' suggests GOF, but no PMID",
            )

    return ("KEEP", "")


# ── 5. Apply QC to main dataframe ─────────────────────────────────────────────
actions, reasons = [], []
for _, row in df.iterrows():
    a, r = classify_qc(row)
    actions.append(a)
    reasons.append(r)

df["qc_round1_action"]  = actions
df["qc_round1_reason"]  = reasons

# ── 6. GROBID context quality annotation ──────────────────────────────────────
df["grobid_context_quality"] = df.apply(
    lambda r: gq_lookup.get(
        (str(r.get("gene", "")), str(r.get("pmid", "")), str(r.get("hgvs_coding", ""))), ""
    ),
    axis=1,
)

# ── 7. Round-1 derived flag columns ───────────────────────────────────────────
def r1_mechanism(row: pd.Series) -> str:
    action   = row["qc_round1_action"]
    existing = str(row.get("curated_mechanism_class", "")).strip()
    if action in ("EXCLUDE_NON_GOF_LOF", "EXCLUDE_NO_GOF_EVIDENCE"):
        return "LOSS_OF_FUNCTION_NON_GOF"
    if action == "EXCLUDE_NONHUMAN_EVIDENCE":
        return "NON_HUMAN_EXCLUDE"
    if action == "EXCLUDE_GENE_EVIDENCE_MISMATCH":
        return "GENE_EVIDENCE_MISMATCH"
    return existing


def r1_strict_gof(row: pd.Series) -> str:
    action   = row["qc_round1_action"]
    existing = str(row.get("strict_gof_status", "")).strip()
    if action.startswith("EXCLUDE"):
        return "EXCLUDED"
    return existing


def r1_include_final(row: pd.Series) -> str:
    action = row["qc_round1_action"]
    if action in ("KEEP", "ADD"):
        return "YES"
    if action.startswith("EXCLUDE"):
        return "NO"
    return "REVIEW"


def r1_include_hn(row: pd.Series) -> str:
    action = row["qc_round1_action"]
    mech   = str(row.get("curated_mechanism_class", "")).strip()
    strict = str(row.get("strict_gof_status", "")).strip()
    if action.startswith("EXCLUDE") or action.startswith("REVIEW"):
        return "NO"
    # KEEP or ADD
    if strict in ("TRUE_GOF",) or mech in ("HYPERMORPH", "HYPERMORPH_TRUE_GOF"):
        return "YES"
    if mech in ("DOMINANT_NEGATIVE_ANTIMORPH", "NEOMORPH_TRUE_GOF", "TOXIC_GOF",
                "NMD_ESCAPE_EFFECT_REVIEW", "NMD_ESCAPE"):
        return "YES_NEOMORPH_DN"
    if strict in ("ANTIMORPH_DN",):
        return "YES_NEOMORPH_DN"
    return "REVIEW"


df["curated_mechanism_class_round1"]    = df.apply(r1_mechanism,    axis=1)
df["strict_gof_status_round1"]          = df.apply(r1_strict_gof,   axis=1)
df["include_in_final_dataset_round1"]   = df.apply(r1_include_final, axis=1)
df["include_in_strict_HN_analysis_round1"] = df.apply(r1_include_hn, axis=1)
df["hn_view_round1"] = df.apply(
    lambda r: (
        "INCLUDE" if r["include_in_strict_HN_analysis_round1"] in ("YES", "YES_NEOMORPH_DN")
        else "EXCLUDE" if r["include_in_final_dataset_round1"] == "NO"
        else "REVIEW"
    ),
    axis=1,
)

# ── 8. ADD rows from Candidate_Additions ──────────────────────────────────────
ADD_CANDIDATES = [
    # 1. CAV1 p.P158PfsX22 — dominant negative
    dict(
        gene="CAV1", hgvs_coding="c.(unknown — tool-detected)", hgvs_protein="p.Pro158PhefsTer22",
        final_variant_type="frameshift", pmid="28904206", confidence="medium",
        gof_effect="Dominant negative form of CAV1 impairs caveolae assembly via new disease mechanism",
        evidence_text=(
            "A mutation that gives rise to a dominant negative form of CAV1, defining a new "
            "mechanism by which disease-associated mutations in CAV1 impair caveolae assembly."
        ),
        source="TOOL_CANDIDATE_ADD", class_="N",
        curated_mechanism_class="DOMINANT_NEGATIVE_ANTIMORPH",
        strict_gof_status="ANTIMORPH_DN",
        qc_round1_action="ADD",
        qc_round1_reason="ADD_CANDIDATE: PubTator3 tool-only review — dominant negative CAV1 frameshift (PMID 28904206)",
        curated_mechanism_class_round1="DOMINANT_NEGATIVE_ANTIMORPH",
        strict_gof_status_round1="ANTIMORPH_DN",
        include_in_final_dataset_round1="YES",
        include_in_strict_HN_analysis_round1="YES_NEOMORPH_DN",
        hn_view_round1="INCLUDE",
    ),
    # 2. KCNH2/hERG L539fs — dominant negative
    dict(
        gene="KCNH2", hgvs_coding="c.(unknown — tool-detected)", hgvs_protein="p.Leu539fsTer?",
        final_variant_type="frameshift", pmid="29752336", confidence="medium",
        gof_effect="Dominant negative suppression of WT-HERG (KCNH2) by truncated hERG",
        evidence_text=(
            "ALLN can restore HERG-A561V mutant protein trafficking and rescue the dominant "
            "negative suppression of WT-HERG."
        ),
        source="TOOL_CANDIDATE_ADD", class_="N",
        curated_mechanism_class="DOMINANT_NEGATIVE_ANTIMORPH",
        strict_gof_status="ANTIMORPH_DN",
        qc_round1_action="ADD",
        qc_round1_reason="ADD_CANDIDATE: PubTator3 tool-only review — dominant negative KCNH2/hERG frameshift (PMID 29752336)",
        curated_mechanism_class_round1="DOMINANT_NEGATIVE_ANTIMORPH",
        strict_gof_status_round1="ANTIMORPH_DN",
        include_in_final_dataset_round1="YES",
        include_in_strict_HN_analysis_round1="YES_NEOMORPH_DN",
        hn_view_round1="INCLUDE",
    ),
    # 3. FBN1 p.Tyr2596Thrfs*86 — dominant negative
    dict(
        gene="FBN1", hgvs_coding="c.(unknown — tool-detected)", hgvs_protein="p.Tyr2596Thrfs*86",
        final_variant_type="frameshift", pmid="31774634", confidence="medium",
        gof_effect="Dominant-negative FBN1 interaction upregulates SMAD2 phosphorylation — MPLS phenotype",
        evidence_text=(
            "Mutations appear to upregulate SMAD2 phosphorylation in vitro. We provide direct "
            "evidence that a dominant-negative interaction of FBN1 potentially explains the complex "
            "MPLS phenotypes."
        ),
        source="TOOL_CANDIDATE_ADD", class_="N",
        curated_mechanism_class="DOMINANT_NEGATIVE_ANTIMORPH",
        strict_gof_status="ANTIMORPH_DN",
        qc_round1_action="ADD",
        qc_round1_reason="ADD_CANDIDATE: PubTator3 tool-only review — dominant negative FBN1 frameshift (PMID 31774634)",
        curated_mechanism_class_round1="DOMINANT_NEGATIVE_ANTIMORPH",
        strict_gof_status_round1="ANTIMORPH_DN",
        include_in_final_dataset_round1="YES",
        include_in_strict_HN_analysis_round1="YES_NEOMORPH_DN",
        hn_view_round1="INCLUDE",
    ),
    # 4. SCNN1B stop@592 — constitutive GOF (Liddle-type)
    dict(
        gene="SCNN1B", hgvs_coding="c.(?C>T)", hgvs_protein="p.Tyr592*",
        final_variant_type="stop_gain", pmid="35774371", confidence="high",
        gof_effect=(
            "Stop at Tyr592 removes C-terminal PY motif → loss of Nedd4-2 ubiquitin ligase "
            "binding → constitutive ENaC surface expression and channel hyperactivity "
            "(Liddle syndrome-type GOF)"
        ),
        evidence_text=(
            "Stop codon at position 592, influencing the crucial PY motif and resulting in "
            "reduced inactivation of the ENaCs. Intra-familial phenotypic heterogeneity observed."
        ),
        source="TOOL_CANDIDATE_ADD", class_="H",
        curated_mechanism_class="HYPERMORPH_TRUE_GOF",
        strict_gof_status="TRUE_GOF",
        qc_round1_action="ADD",
        qc_round1_reason=(
            "ADD_CANDIDATE: PubTator3 tool-only review — SCNN1B stop@PY motif. "
            "Classic Liddle syndrome mechanism: truncation of PY motif prevents Nedd4-2-mediated "
            "internalization → constitutive ENaC activity. Strong GOF. (PMID 35774371)"
        ),
        curated_mechanism_class_round1="HYPERMORPH_TRUE_GOF",
        strict_gof_status_round1="TRUE_GOF",
        include_in_final_dataset_round1="YES",
        include_in_strict_HN_analysis_round1="YES",
        hn_view_round1="INCLUDE",
    ),
    # 5. CXCR4 V340fs — explicit GOF (WHIM syndrome)
    dict(
        gene="CXCR4", hgvs_coding="c.(V340fs)", hgvs_protein="p.Val340fsTer?",
        final_variant_type="frameshift", pmid="39877344", confidence="medium",
        gof_effect=(
            "CXCR4 V340fs gain-of-function: C-terminal truncation prevents receptor "
            "desensitization/internalization → prolonged CXCL12 signalling → WHIM syndrome"
        ),
        evidence_text=(
            "Heterogeneous phenotype of a Chinese Familial WHIM syndrome with CXCR4V340fs "
            "gain-of-function mutation."
        ),
        source="TOOL_CANDIDATE_ADD", class_="H",
        curated_mechanism_class="HYPERMORPH_TRUE_GOF",
        strict_gof_status="TRUE_GOF_REVIEW",
        qc_round1_action="ADD",
        qc_round1_reason=(
            "ADD_CANDIDATE: PubTator3 tool-only review — CXCR4 V340fs explicitly labeled "
            "'gain-of-function' in paper title; WHIM syndrome. Existing dataset has V340Lfs*27 "
            "(PMID 39575248); this is different PMID, same variant region. (PMID 39877344)"
        ),
        curated_mechanism_class_round1="HYPERMORPH_TRUE_GOF",
        strict_gof_status_round1="TRUE_GOF_REVIEW",
        include_in_final_dataset_round1="YES",
        include_in_strict_HN_analysis_round1="REVIEW",
        hn_view_round1="REVIEW",
    ),
    # 6. Unknown gene (Entrez 54894) R117fs — REVIEW_GENE_ID_REQUIRED
    dict(
        gene="UNKNOWN_ENTREZ54894", hgvs_coding="c.(R117fs)", hgvs_protein="p.Arg117fsTer?",
        final_variant_type="frameshift", pmid="34214079", confidence="low",
        gof_effect=(
            "Putative GOF via RAS/MAPK pathway activation — gene identity requires "
            "confirmation (Entrez ID 54894 not mapped)"
        ),
        evidence_text=(
            "Transcriptional profiles similar to BRAF missense mutations with activated "
            "RAS/MAPK signaling, consistent with KRAS signaling pathways being GOF in both "
            "R117fs and G659fs variants."
        ),
        source="TOOL_CANDIDATE_REVIEW", class_="",
        curated_mechanism_class="",
        strict_gof_status="",
        qc_round1_action="REVIEW_GENE_ID_REQUIRED",
        qc_round1_reason=(
            "Candidate ADD but Entrez gene ID 54894 could not be confirmed. "
            "Verify gene symbol at https://www.ncbi.nlm.nih.gov/gene/54894 before adding. "
            "(PMID 34214079)"
        ),
        curated_mechanism_class_round1="",
        strict_gof_status_round1="",
        include_in_final_dataset_round1="REVIEW",
        include_in_strict_HN_analysis_round1="NO",
        hn_view_round1="REVIEW",
    ),
]

# Build ADD dataframe — fill missing columns with ""
add_records = []
for rec in ADD_CANDIDATES:
    rec2 = {k: v for k, v in rec.items() if k != "class_"}
    if "class_" in rec:
        rec2["class"] = rec["class_"]
    add_records.append(rec2)

df_add = pd.DataFrame(add_records)
for col in df.columns:
    if col not in df_add.columns:
        df_add[col] = ""
df_add = df_add.reindex(columns=df.columns, fill_value="")

# Check for CXCR4 V340 existing entries
cxcr4_check = df[
    (df["gene"] == "CXCR4") &
    (df["hgvs_protein"].str.contains(r"V340|Val340", na=False, regex=True))
]
if len(cxcr4_check):
    print(f"Note: CXCR4 V340 already in dataset: {len(cxcr4_check)} row(s) — new PMID 39877344 is different paper")

df_full = pd.concat([df, df_add], ignore_index=True)
print(f"After ADD: {len(df_full)} rows total ({len(df_add)} new)")

# ── 9. Print action counts ─────────────────────────────────────────────────────
print(f"\n{'='*65}")
print("=== QC Round 1 — Action Distribution ===")
print(f"{'='*65}")
for action, cnt in df_full["qc_round1_action"].value_counts().items():
    print(f"  {action:<45} {cnt:>4}")
print(f"{'='*65}")
print(f"  include_in_final_dataset_round1 = YES:    "
      f"{(df_full['include_in_final_dataset_round1']=='YES').sum()}")
print(f"  include_in_final_dataset_round1 = NO:     "
      f"{(df_full['include_in_final_dataset_round1']=='NO').sum()}")
print(f"  include_in_final_dataset_round1 = REVIEW: "
      f"{(df_full['include_in_final_dataset_round1']=='REVIEW').sum()}")

# ── 10. Build output subsets ───────────────────────────────────────────────────
mask_add      = df_full["qc_round1_action"] == "ADD"
mask_excl     = df_full["qc_round1_action"].str.startswith("EXCLUDE")
mask_rlof     = df_full["curated_mechanism_class_round1"] == "LOSS_OF_FUNCTION_NON_GOF"
mask_rev      = df_full["qc_round1_action"].str.startswith("REVIEW")
mask_opa1     = (df_full["gene"] == "OPA1") & (df_full["qc_round1_action"] == "REVIEW_POSSIBLE_LOF_HAPLOINSUFFICIENCY")
mask_pmid_mis = df_full["qc_round1_action"] == "REVIEW_PMID_MISSING_POSSIBLE_GOF"

# ── 11. Column ordering ────────────────────────────────────────────────────────
QC_COLS = [
    "qc_round1_action", "qc_round1_reason",
    "include_in_final_dataset_round1", "include_in_strict_HN_analysis_round1", "hn_view_round1",
    "curated_mechanism_class_round1", "strict_gof_status_round1",
    "grobid_context_quality",
]
ORIG_PRIORITY = [
    "gene", "hgvs_coding", "hgvs_protein", "final_variant_type", "confidence",
    "pmid", "doi", "year", "title", "source",
    "gof_effect", "evidence_text",
    "curated_mechanism_class", "strict_gof_status", "hn_view",
    "class", "functional_category",
    "transcript_id", "hgvs_full", "hgvs_genomic", "spdi", "chromosome", "position_hg38",
    "disease", "omim", "hgvs_variant_type", "is_human", "is_ptc", "reason",
    "variant_type", "note", "needs_review",
    "curation_reason", "evidence_strength", "dedup_key",
    "model_or_nonhuman_evidence_cue", "evidence_species_inference",
    "conflict_group_id", "review_priority",
]
col_order = QC_COLS + [c for c in ORIG_PRIORITY if c in df_full.columns] + \
            [c for c in df_full.columns if c not in QC_COLS and c not in ORIG_PRIORITY]
df_full = df_full[[c for c in col_order if c in df_full.columns]]

# ── 12. Summary sheet ─────────────────────────────────────────────────────────
summary_rows = [
    ["=== GOF Curation QC Round 1 Summary ===", ""],
    ["Date", "2026-04-30"],
    ["", ""],
    ["[ Input ]", ""],
    ["Original Curated_All rows",                   len(df)],
    ["ADD rows (new from Candidate_Additions)",      mask_add.sum()],
    ["Total rows in output",                        len(df_full)],
    ["", ""],
    ["[ qc_round1_action Distribution ]", ""],
]
for action, cnt in df_full["qc_round1_action"].value_counts().items():
    summary_rows.append([f"  {action}", cnt])
summary_rows += [
    ["", ""],
    ["[ Final inclusion flags ]", ""],
    ["include_in_final_dataset_round1 = YES",    (df_full["include_in_final_dataset_round1"]=="YES").sum()],
    ["include_in_final_dataset_round1 = NO",     (df_full["include_in_final_dataset_round1"]=="NO").sum()],
    ["include_in_final_dataset_round1 = REVIEW", (df_full["include_in_final_dataset_round1"]=="REVIEW").sum()],
    ["", ""],
    ["include_in_strict_HN_analysis_round1 = YES",           (df_full["include_in_strict_HN_analysis_round1"]=="YES").sum()],
    ["include_in_strict_HN_analysis_round1 = YES_NEOMORPH_DN",(df_full["include_in_strict_HN_analysis_round1"]=="YES_NEOMORPH_DN").sum()],
    ["include_in_strict_HN_analysis_round1 = REVIEW",        (df_full["include_in_strict_HN_analysis_round1"]=="REVIEW").sum()],
    ["include_in_strict_HN_analysis_round1 = NO",            (df_full["include_in_strict_HN_analysis_round1"]=="NO").sum()],
    ["", ""],
    ["[ Excluded rows by type ]", ""],
    ["EXCLUDE_NONHUMAN_EVIDENCE",       (df_full["qc_round1_action"]=="EXCLUDE_NONHUMAN_EVIDENCE").sum()],
    ["EXCLUDE_NON_GOF_LOF",            (df_full["qc_round1_action"]=="EXCLUDE_NON_GOF_LOF").sum()],
    ["EXCLUDE_NO_GOF_EVIDENCE",        (df_full["qc_round1_action"]=="EXCLUDE_NO_GOF_EVIDENCE").sum()],
    ["EXCLUDE_GENE_EVIDENCE_MISMATCH", (df_full["qc_round1_action"]=="EXCLUDE_GENE_EVIDENCE_MISMATCH").sum()],
    ["", ""],
    ["[ GROBID context quality (annotated rows) ]", ""],
]
for qval, cnt in df_full["grobid_context_quality"].value_counts().items():
    if qval:
        summary_rows.append([f"  grobid_context_quality={qval}", cnt])

df_summary = pd.DataFrame(summary_rows, columns=["항목", "값"])

# ── 13. Rules sheet ────────────────────────────────────────────────────────────
RULES = [
    ["Action", "Definition", "Conservative?", "Affected genes/cases"],
    ["KEEP",
     "Passes all QC checks — include in final dataset",
     "Yes — default for all ambiguous rows",
     "Majority of 2,480 rows"],
    ["ADD",
     "New candidate from tool-only review, manually confirmed with GOF evidence",
     "Yes — only 5 clear cases added",
     "CAV1 p.P158PfsX22, KCNH2 L539fs, FBN1 p.Tyr2596Thrfs*86, SCNN1B p.Tyr592*, CXCR4 V340fs"],
    ["EXCLUDE_NONHUMAN_EVIDENCE",
     "Evidence/title text clearly refers to a non-human organism",
     "Conservative: only when non-human keywords unambiguous",
     "DNAJC21 (GS3/rice, PMID 23641184), NLD (maize), ATF2 (dCREB2/Drosophila), REC8 (Daphnia-like)"],
    ["EXCLUDE_NON_GOF_LOF",
     "Evidence exclusively describes LOF/functional reduction, no GOF override keyword",
     "Conservative: specific gene+PMID combinations only",
     "EMD (all, X-linked LOF disease), BSG, CFHR5, TBX6, CLCN2, LRRC8A, KCNJ1(10532965), KCNJ1(NOT_GOF)"],
    ["EXCLUDE_NO_GOF_EVIDENCE",
     "Explicit 'no gain-of-function effect demonstrated' statement in gof_effect or evidence_text",
     "Conservative: verbatim negation only",
     "SMPD1 (PMID 22558155, 3 rows)"],
    ["EXCLUDE_GENE_EVIDENCE_MISMATCH",
     "Gene field and evidence text clearly refer to different genes",
     "Conservative: only clear mismatches",
     "LOXL2 rows where evidence describes loricrin (LOR gene)"],
    ["REVIEW_POSSIBLE_LOF_HAPLOINSUFFICIENCY",
     "OPA1 PMID 14961560 — from variant catalogue table; no GOF cue found in fulltext GROBID scan",
     "Kept as REVIEW — not auto-excluded",
     "OPA1 x6 rows from PMID 14961560"],
    ["REVIEW_PMID_MISSING_POSSIBLE_GOF",
     "No PMID but plausible GOF mechanism — needs citation before final inclusion",
     "Kept as REVIEW — possible TRUE_GOF",
     "KAT6A (4 variants), ALAS2 c.1670_1671TC>GA"],
    ["REVIEW_GENE_ID_REQUIRED",
     "Candidate ADD but gene identity (Entrez 54894) not confirmed",
     "Kept as REVIEW — not added until gene verified",
     "R117fs PMID 34214079 (Entrez 54894)"],
    ["REVIEW_MECHANISM",
     "Conflicting signals: explicit no-GOF statement AND GOF override keyword co-present",
     "Kept as REVIEW — manual decision required",
     "Rare — typically complex papers with multiple variant types"],
    ["", "", "", ""],
    ["Column", "Meaning", "", ""],
    ["include_in_final_dataset_round1",
     "YES=include, NO=exclude, REVIEW=needs manual decision", "", ""],
    ["include_in_strict_HN_analysis_round1",
     "YES=strict GOF (hypermorph), YES_NEOMORPH_DN=neomorph/DN included, "
     "NO=excluded/review, REVIEW=needs decision", "", ""],
    ["grobid_context_quality",
     "GOOD=variant+GOF cue in same biological sentence, "
     "BAD_SECTION=snippet in affiliation/methods/reagents, "
     "VARIANT_ONLY_NO_GOF=variant found but no GOF cue", "", ""],
    ["", "", "", ""],
    ["Important notes", "", "", ""],
    ["1. No original rows deleted — use include_in_final_dataset_round1 as filter", "", "", ""],
    ["2. KEEP rows with existing strict_gof_status=NOT_GOF/QC_FAIL are preserved with their original labels", "", "", ""],
    ["3. REVIEW rows require human curator decision before final dataset freezing", "", "", ""],
    ["4. LOXL2 c.709dupC (no PMID, strict=TRUE_GOF) was preserved as KEEP — lacks loricrin keyword", "", "", ""],
]
df_rules = pd.DataFrame(RULES[1:], columns=RULES[0])

# ── 14. Write Excel ────────────────────────────────────────────────────────────
ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

def clean_df(d: pd.DataFrame) -> pd.DataFrame:
    return d.map(lambda v: ILLEGAL_CHARS_RE.sub("", v) if isinstance(v, str) else v)

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)

ACTION_FILLS = {
    "KEEP":                              PatternFill("solid", fgColor="E2EFDA"),
    "ADD":                               PatternFill("solid", fgColor="C6EFCE"),
    "EXCLUDE_NONHUMAN_EVIDENCE":         PatternFill("solid", fgColor="F4CCCC"),
    "EXCLUDE_NON_GOF_LOF":              PatternFill("solid", fgColor="F4CCCC"),
    "EXCLUDE_NO_GOF_EVIDENCE":          PatternFill("solid", fgColor="F4CCCC"),
    "EXCLUDE_GENE_EVIDENCE_MISMATCH":   PatternFill("solid", fgColor="F4CCCC"),
    "REVIEW_POSSIBLE_LOF_HAPLOINSUFFICIENCY": PatternFill("solid", fgColor="FFF2CC"),
    "REVIEW_PMID_MISSING_POSSIBLE_GOF": PatternFill("solid", fgColor="FFF2CC"),
    "REVIEW_GENE_ID_REQUIRED":          PatternFill("solid", fgColor="FFF2CC"),
    "REVIEW_MECHANISM":                 PatternFill("solid", fgColor="FCE5CD"),
    "REVIEW_CONTEXT_QUALITY":           PatternFill("solid", fgColor="FFF2CC"),
}

def style_ws(ws, df_s: pd.DataFrame = None, action_col: str = None):
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=False)
    if action_col and df_s is not None and action_col in df_s.columns:
        col_idx = list(df_s.columns).index(action_col) + 1
        for r_idx, (_, row) in enumerate(df_s.iterrows(), start=2):
            val  = str(row.get(action_col, ""))
            fill = ACTION_FILLS.get(val)
            if fill:
                for cell in ws[r_idx]:
                    cell.fill = fill
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0 for c in col_cells), default=0)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 60)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


print(f"\nWriting → {OUT_XL}")
with pd.ExcelWriter(OUT_XL, engine="openpyxl") as writer:
    # 1. Curated_All_With_QC_Action
    clean_df(df_full).to_excel(writer, sheet_name="Curated_All_With_QC_Action", index=False)
    style_ws(writer.sheets["Curated_All_With_QC_Action"], df_full, action_col="qc_round1_action")

    # 2. Rows_To_Add
    df_add_out = df_full[mask_add].copy()
    clean_df(df_add_out).to_excel(writer, sheet_name="Rows_To_Add", index=False)
    style_ws(writer.sheets["Rows_To_Add"])

    # 3. Rows_To_Exclude
    df_excl = df_full[mask_excl].copy()
    clean_df(df_excl).to_excel(writer, sheet_name="Rows_To_Exclude", index=False)
    style_ws(writer.sheets["Rows_To_Exclude"], df_excl, action_col="qc_round1_action")

    # 4. Rows_To_Reclassify_NonGOF
    df_rlof = df_full[mask_rlof].copy()
    clean_df(df_rlof).to_excel(writer, sheet_name="Rows_To_Reclassify_NonGOF", index=False)
    style_ws(writer.sheets["Rows_To_Reclassify_NonGOF"], df_rlof, action_col="qc_round1_action")

    # 5. Rows_To_Review
    df_rev = df_full[mask_rev].copy()
    clean_df(df_rev).to_excel(writer, sheet_name="Rows_To_Review", index=False)
    style_ws(writer.sheets["Rows_To_Review"], df_rev, action_col="qc_round1_action")

    # 6. OPA1_Review
    df_opa1 = df_full[mask_opa1].copy()
    clean_df(df_opa1).to_excel(writer, sheet_name="OPA1_Review", index=False)
    style_ws(writer.sheets["OPA1_Review"])

    # 7. PMID_Missing_Review
    df_pmid = df_full[mask_pmid_mis].copy()
    clean_df(df_pmid).to_excel(writer, sheet_name="PMID_Missing_Review", index=False)
    style_ws(writer.sheets["PMID_Missing_Review"])

    # 8. Summary
    df_summary.to_excel(writer, sheet_name="Summary", index=False)
    style_ws(writer.sheets["Summary"])

    # 9. Rules
    df_rules.to_excel(writer, sheet_name="Rules", index=False)
    style_ws(writer.sheets["Rules"])

print(f"Saved: {OUT_XL}")
print(f"\n{'='*65}")
print("=== Final check ===")
print(f"  Curated_All_With_QC_Action : {len(df_full)} rows")
print(f"  Rows_To_Add                : {len(df_add_out)} rows")
print(f"  Rows_To_Exclude            : {len(df_excl)} rows")
print(f"  Rows_To_Reclassify_NonGOF  : {len(df_rlof)} rows")
print(f"  Rows_To_Review             : {len(df_rev)} rows")
print(f"  OPA1_Review                : {len(df_opa1)} rows")
print(f"  PMID_Missing_Review        : {len(df_pmid)} rows")
print(f"{'='*65}")
